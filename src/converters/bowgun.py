from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.data.text_db import TextDB, TextSource
from src.data.user3 import load_user3_table
from src.pipeline.transforms import transform_workbook
from src.utils.log import file_size, info

Table = list[dict]

ZH_HANS_LANGUAGE_ID = 13

CUSTOMIZE_DATA_PATH = "STM/GameDesign/Common/Equip/BowgunCustomizeData.user.3.json"
CUSTOMIZE_ITEM_PATH = "STM/GameDesign/Common/Equip/BowgunCustomizeItemData.user.3.json"
HEAVY_BOWGUN_PATH = "STM/GameDesign/Common/Weapon/HeavyBowgun.user.3.json"
LIGHT_BOWGUN_PATH = "STM/GameDesign/Common/Weapon/LightBowgun.user.3.json"

SHELL_NAMES = [
    "通常弹",
    "贯通弹",
    "散弹",
    "穿甲榴弹",
    "扩散弹",
    "斩裂弹",
    "龙击弹",
    "火炎弹",
    "水冷弹",
    "电击弹",
    "冰结弹",
    "灭龙弹",
    "毒弹",
    "麻痹弹",
    "眠弹",
]

SHELL_COLORS = {
    "通常弹": "D9D9D9",
    "贯通弹": "C5D9F1",
    "散弹": "EBF1DE",
    "穿甲榴弹": "DDD9C4",
    "扩散弹": "F2DCDB",
    "斩裂弹": "FFFFFF",
    "龙击弹": "FDE9D9",
    "火炎弹": "FABB94",
    "水冷弹": "8DB4E2",
    "电击弹": "EFF991",
    "冰结弹": "B1EBF1",
    "灭龙弹": "EFCDCD",
    "毒弹": "F4DDFF",
    "麻痹弹": "FBF4E1",
    "眠弹": "DEFCFE",
}

SHELL_LEVELS = {
    "SL_000": "1",
    "SL_001": "2",
    "SL_002": "3",
    "NONE": "",
}

ENERGY_EFFICIENCY = {
    "EET_000": "Lv.1",
    "EET_001": "Lv.2",
}

AMMO_STRENGTH = {
    "AST_000": "一般／龙热射击基本型",
    "AST_001": "一般模式特化型Ⅰ",
    "AST_002": "一般模式特化型Ⅱ",
    "AST_003": "龙热模式特化型Ⅰ",
    "AST_004": "龙热模式特化型Ⅱ",
}

ENERGY_SHELL_NORMAL = {
    "ESN_000": "龙热机关龙弹",
    "ESN_001": "龙热穿甲弹",
}

ENERGY_SHELL_POWER = {
    "ESP_000": "龙热抵消弹",
    "ESP_001": "龙热榴弹",
}

SPECIAL_AMMO = {
    "SET_BOMB": "起爆龙弹",
    "CHATCH": "附着龙弹",
}

CUSTOMIZE_HEADERS = [
    "PatternID",
    "Index",
    "DataID",
    "ItemID",
    "IsFirstSetting",
    "Value_1",
    "Value_2",
    "Name",
    "Explain",
]

HEAVY_HEADERS = [
    "武器ID",
    "名称",
    "模型ID",
    "稀有度",
    "攻击",
    "防御",
    "会心率",
    "孔位",
    "Main Shell",
    *SHELL_NAMES,
    "能量弹种类(Normal)",
    "能量弹种类(Power)",
    "龙热槽回复效率",
    "弹药模式",
    "客制零件槽1（多选一）",
    "客制零件槽2（多选一）",
    "技能1",
    "技能2",
    "技能3",
    "技能4",
    "描述",
]

LIGHT_HEADERS = [
    "武器ID",
    "名称",
    "模型ID",
    "稀有度",
    "攻击",
    "防御",
    "会心率",
    "孔位",
    "Main Shell",
    *SHELL_NAMES,
    "特殊弹药",
    "客制零件槽1（多选一）",
    "客制零件槽2（多选一）",
    "技能1",
    "技能2",
    "技能3",
    "技能4",
    "描述",
]

RARE_RE = re.compile(r"^RARE(\d+)$")
HAN_RE = re.compile(r"[\u4e00-\u9fff]")
INVALID_SHEET_CHARS = str.maketrans({c: "_" for c in r'[]:*?/\\'})


def export_bowgun_workbooks(output_dir: Path, natives_dir: Path, text_source: TextSource) -> None:
    text_db = text_source.build(ZH_HANS_LANGUAGE_ID)

    info("    Loading bowgun customize tables")
    customize_rows = _customize_rows(natives_dir, text_db)
    customize_path = output_dir / "Bowgun_Custom.xlsx"
    _write_table(customize_path, "弩枪客制零件信息", CUSTOMIZE_HEADERS, customize_rows)
    info(f"    Saved workbook: {customize_path} ({file_size(customize_path)})")

    custom_items = _custom_items_by_pattern(customize_rows)
    info("    Loading bowgun weapon tables")
    sheets = {
        "Wp_HeavyBowgun": _load(natives_dir, HEAVY_BOWGUN_PATH, text_db) or [],
        "Wp_LightBowgun": _load(natives_dir, LIGHT_BOWGUN_PATH, text_db) or [],
    }
    sheets = transform_workbook(
        "EquipCollection.xlsx",
        sheets,
        lambda relative_path: _load(natives_dir, relative_path, text_db),
    )

    heavy_rows = [_heavy_bowgun_row(row, custom_items) for row in sheets.get("Wp_HeavyBowgun", [])]
    heavy_path = output_dir / "HeavyBowgun.xlsx"
    _write_table(heavy_path, "HeavyBowgun", HEAVY_HEADERS, heavy_rows, SHELL_NAMES)
    info(f"    Saved workbook: {heavy_path} ({file_size(heavy_path)})")

    light_rows = [_light_bowgun_row(row, custom_items) for row in sheets.get("Wp_LightBowgun", [])]
    light_path = output_dir / "LightBowgun.xlsx"
    _write_table(light_path, "LightBowgun", LIGHT_HEADERS, light_rows, SHELL_NAMES)
    info(f"    Saved workbook: {light_path} ({file_size(light_path)})")


def _customize_rows(natives_dir: Path, text_db: TextDB) -> Table:
    patterns = _load(natives_dir, CUSTOMIZE_DATA_PATH, text_db) or []
    items = {
        row.get("ItemID"): row
        for row in (_load(natives_dir, CUSTOMIZE_ITEM_PATH, text_db) or [])
    }

    rows = []
    for pattern in patterns:
        item = items.get(pattern.get("ItemID"), {})
        values = _as_list(item.get("Value"))
        rows.append(
            {
                "PatternID": pattern.get("PatternID"),
                "Index": pattern.get("Index"),
                "DataID": pattern.get("DataId"),
                "ItemID": pattern.get("ItemID"),
                "IsFirstSetting": pattern.get("IsFirstSetting"),
                "Value_1": _at(values, 0),
                "Value_2": _at(values, 1),
                "Name": item.get("Name", ""),
                "Explain": item.get("Explain", ""),
            }
        )
    return sorted(rows, key=lambda row: (_id_key(row.get("PatternID")), row.get("Index") or 0))


def _heavy_bowgun_row(row: dict, custom_items: dict[str, list[str]]) -> dict:
    patterns = _patterns(row.get("CustomizePattern"))
    skills = _skills(row.get("SkillAndLevel"))
    shell_values = _shell_values(row.get("ShellLv"), row.get("ShellNum"))
    return _ordered_row(
        HEAVY_HEADERS,
        [
            row.get("Index"),
            row.get("Name"),
            row.get("ModelId"),
            _rare(row.get("Rare")),
            row.get("Attack"),
            row.get("Defense"),
            row.get("Critical"),
            _slot(row.get("SlotLevel")),
            row.get("MainShell"),
            *shell_values,
            _mapped(row.get("EnergyShellTypeNormal"), ENERGY_SHELL_NORMAL),
            _mapped(row.get("EnergyShellTypePower"), ENERGY_SHELL_POWER),
            _mapped(row.get("EnergyEfficiency"), ENERGY_EFFICIENCY),
            _mapped(row.get("AmmoStrength"), AMMO_STRENGTH),
            _custom_items(custom_items, _at(patterns, 0)),
            _custom_items(custom_items, _at(patterns, 1)),
            *skills,
            row.get("Explain"),
        ],
    )


def _light_bowgun_row(row: dict, custom_items: dict[str, list[str]]) -> dict:
    patterns = _patterns(row.get("CustomizePattern"))
    skills = _skills(row.get("SkillAndLevel"))
    shell_values = _shell_values(row.get("ShellLv"), row.get("ShellNum"), row.get("IsRappid"))
    return _ordered_row(
        LIGHT_HEADERS,
        [
            row.get("Index"),
            row.get("Name"),
            row.get("ModelId"),
            _rare(row.get("Rare")),
            row.get("Attack"),
            row.get("Defense"),
            row.get("Critical"),
            _slot(row.get("SlotLevel")),
            row.get("MainShell"),
            *shell_values,
            _mapped(row.get("Wp13SpecialAmmo"), SPECIAL_AMMO),
            _custom_items(custom_items, _at(patterns, 0)),
            _custom_items(custom_items, _at(patterns, 1)),
            *skills,
            row.get("Explain"),
        ],
    )


def _load(natives_dir: Path, relative_path: str, text_db: TextDB) -> Table | None:
    source = natives_dir / relative_path
    if not source.exists():
        info(f"    Skip missing: {source}")
        return None
    frame = load_user3_table(source, text_db)
    info(f"    Loaded {relative_path}: {len(frame)} row(s)")
    return frame


def _custom_items_by_pattern(rows: Table) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = {}
    for row in rows:
        pattern_id = row.get("PatternID")
        name = row.get("Name")
        if pattern_id and name:
            grouped.setdefault(pattern_id, []).append(str(name))
    return grouped


def _custom_items(custom_items: dict[str, list[str]], pattern: Any) -> str:
    names = custom_items.get(str(pattern), []) if pattern else []
    return f"[{', '.join(names)}]"


def _shell_values(levels: Any, nums: Any, rapid_flags: Any = None) -> list[str]:
    level_list = _as_list(levels)
    num_list = _as_list(nums)
    rapid_list = _as_list(rapid_flags)
    values = []
    for index in range(len(SHELL_NAMES)):
        level = _at(level_list, index)
        num = _int(_at(num_list, index))
        if num == 0:
            values.append("")
            continue
        suffix = " [速射]" if bool(_at(rapid_list, index)) else ""
        if level == "NONE":
            values.append(f"Lv.-: {num}发{suffix}")
        else:
            values.append(f"Lv.{SHELL_LEVELS.get(str(level), level)}: {num}发{suffix}")
    return values


def _skills(value: Any) -> list[str]:
    values = [str(item) for item in _as_list(value) if item]
    return (values + ["", "", "", ""])[:4]


def _patterns(value: Any) -> list[str]:
    values = [str(item) for item in _as_list(value) if item]
    return (values + ["", ""])[:2]


def _slot(value: Any) -> str:
    return "-".join(str(item) for item in _as_list(value))


def _rare(value: Any) -> Any:
    if isinstance(value, str):
        match = RARE_RE.match(value)
        if match:
            return int(match.group(1)) + 1
    return value


def _mapped(value: Any, mapping: dict[str, str]) -> Any:
    return mapping.get(value, value)


def _ordered_row(headers: list[str], values: list[Any]) -> dict:
    return {header: value for header, value in zip(headers, values)}


def _write_table(
    path: Path,
    sheet_name: str,
    headers: list[str],
    rows: Table,
    colored_columns: list[str] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_name(sheet_name)
    sheet.append(headers)
    for row in rows:
        sheet.append([_cell(row.get(header)) for header in headers])
    _style_table(sheet, headers, colored_columns or [])
    workbook.save(path)


def _style_table(sheet, headers: list[str], colored_columns: list[str]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    colored_indexes = {
        index
        for index, header in enumerate(headers, start=1)
        if header in set(colored_columns)
    }

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column not in colored_indexes)
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill

    for column_index in colored_indexes:
        color = SHELL_COLORS.get(sheet.cell(1, column_index).value)
        if not color:
            continue
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for row_index in range(1, sheet.max_row + 1):
            sheet.cell(row_index, column_index).fill = fill

    for column_index, header in enumerate(headers, start=1):
        max_width = 80.0 if header in {"描述", "Explain"} else 50.0
        width = max(8.0, _text_width(header) + 2)
        for row_index in range(2, sheet.max_row + 1):
            width = max(width, _text_width(sheet.cell(row_index, column_index).value) + 2)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(width, max_width)


def _cell(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return str(value)
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def _sheet_name(name: str) -> str:
    return name.translate(INVALID_SHEET_CHARS)[:31] or "Sheet"


def _text_width(value: Any) -> float:
    if value is None:
        return 0.0
    total = 0.0
    for char in str(value):
        if char.isascii() and char.isalnum():
            total += 1.2
        elif HAN_RE.search(char):
            total += 2.0
        else:
            total += 1.0
    return total


def _as_list(value: Any) -> list:
    if isinstance(value, list):
        return value
    return [] if value is None or value == "" else [value]


def _at(values: list, index: int) -> Any:
    return values[index] if index < len(values) else ""


def _int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _id_key(value: Any) -> tuple[str, int | str]:
    text = str(value or "")
    if "_" in text:
        prefix, suffix = text.rsplit("_", 1)
        if suffix.isdigit():
            return prefix, int(suffix)
    return text, text
