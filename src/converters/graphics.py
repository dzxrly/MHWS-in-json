import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.data.user3 import load_user3_table, normalize
from src.utils.log import file_size, info

GRAPHICS_PATH = "STM/System/SystemSetting/GraphicsPreset.user.3.json"
NAMES_PATH = Path(__file__).resolve().parent / "data" / "graphics_names_zh_hans.json"
PLATFORM = {
    1: "PS5",
    2: "PS5_Trinity",
    3: "XBSX",
    4: "XBSS",
    5: "PC",
    13: "SteamDeck",
    14: "NSW2",
    15: "NSW2_HANDHELD",
}
USAGE = {
    0: "Default",
    99: "PC_Lowest",
    100: "PC_Low",
    101: "PC_Middle",
    102: "PC_High",
    103: "PC_Highest",
    241: "FastLoadMode",
    242: "DialoguePlaying",
    243: "Title",
}


def export_graphic_preset(output_dir: Path, natives_dir: Path) -> None:
    info(f"    Loading graphic preset: {GRAPHICS_PATH}")
    root = _root(natives_dir / GRAPHICS_PATH)
    rows = load_user3_table(natives_dir / GRAPHICS_PATH)
    presets = [_flatten(_expand(row, root)) for row in rows]
    names = _names()
    attrs = sorted({key for row in presets for key in row} - {"Platform", "Usage", "MaxFPS"})
    info(f"    Graphic presets: {len(presets)} preset(s), {len(attrs)} setting column(s)")

    wb = Workbook()
    wb.remove(wb.active)
    _create_sheet(wb, "基础设置", presets, attrs, names, filtered=True)
    _create_sheet(wb, "未经过Usage筛选", presets, attrs, names, filtered=False)
    path = output_dir / "graphic_preset.xlsx"
    wb.save(path)
    info(f"    Saved workbook: {path} ({file_size(path)})")


def _create_sheet(
    wb: Workbook,
    title: str,
    presets: list[dict],
    attrs: list[str],
    names: dict[str, str],
    filtered: bool,
) -> None:
    ws = wb.create_sheet(title)
    ws.cell(1, 1, "中文名")
    ws.cell(1, 2, "变量名")
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
    columns = _preset_columns(ws, presets, filtered)
    for attr in attrs:
        row_idx = ws.max_row + 1
        ws.cell(row_idx, 1, _cn_name(attr, names))
        ws.cell(row_idx, 2, attr)
        values = [preset.get(attr) for _, preset in columns]
        for col_idx, value in zip((col for col, _ in columns), values):
            ws.cell(row_idx, col_idx, _cell_value(value))
        _color_row(ws, row_idx, values)
    _style(ws)


def _root(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    for item in data:
        if "app.user_data.AppGraphicsSettingPreset" in item:
            return item["app.user_data.AppGraphicsSettingPreset"]
    return {}


def _names() -> dict[str, str]:
    with NAMES_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def _cn_name(attr: str, names: dict[str, str]) -> str:
    variants = [
        attr,
        f"_{attr}",
        attr.replace("StreamingMeshLimit_", "StreamingMeshLimit__"),
        f"_{attr.replace('StreamingMeshLimit_', 'StreamingMeshLimit__')}",
    ]
    for key in variants:
        if key in names:
            return names[key]
    return attr


def _expand(row: dict, root: dict) -> dict:
    out = dict(row)
    _expand_nested(out, "RayTracing", "RayTracing")
    _expand_nested(out, "ExperimentalRayTrace", "ExperimentalRayTrace")
    _expand_match(out, "GIQuality", _wrapped(root.get("_GIQualitySettings")), "Type", "GIQuality")
    _expand_match(out, "DynamicResolutionMode", _wrapped(root.get("_DynamicResolutionParamList")), "Mode", "DynamicResolution")
    _expand_match(out, "StreamingTextureQuality", _wrapped(root.get("_StreamingTextureSettingList")), "Quality", "StreamingTexture")
    _expand_match(
        out,
        "StreamingMeshMinimumLOD",
        _wrapped(root.get("_StreamingMeshLimitList")),
        "StreamingMeshMinimumLodLimit",
        "StreamingMeshLimit",
    )
    return out


def _wrapped(rows: list | None) -> list[dict]:
    return [_unwrap(row) for row in rows or []]


def _unwrap(row: Any) -> dict:
    if isinstance(row, dict) and len(row) == 1:
        value = next(iter(row.values()))
        if isinstance(value, dict):
            return normalize(value)
    value = normalize(row)
    return value if isinstance(value, dict) else {}


def _expand_nested(row: dict, key: str, prefix: str) -> None:
    value = row.pop(key, None)
    if not isinstance(value, dict):
        return
    nested_values = value.values() if any(isinstance(v, dict) for v in value.values()) else [value]
    for nested in nested_values:
        if not isinstance(nested, dict):
            continue
        for sub_key, sub_value in nested.items():
            row[f"{prefix}_{sub_key}"] = _cell_value(sub_value)


def _expand_match(row: dict, source_key: str, rows: list[dict], match_key: str, prefix: str) -> None:
    source_value = row.get(source_key)
    for settings in rows:
        if settings.get(match_key) != source_value:
            continue
        for key, value in settings.items():
            if key != match_key:
                row[f"{prefix}_{key}"] = _cell_value(value)
        return


def _preset_columns(ws, presets: list[dict], filtered: bool) -> list[tuple[int, dict]]:
    columns = []
    col_idx = 3
    for platform in sorted({row.get("Platform") for row in presets}, key=lambda v: (v is None, v)):
        rows = [row for row in presets if row.get("Platform") == platform]
        rows = [row for row in rows if (not filtered or _include(row))]
        rows.sort(key=lambda row: (row.get("MaxFPS") or 0, row.get("Usage") or 0))
        if not rows:
            continue
        start = col_idx
        fps_groups: dict[Any, list[int]] = {}
        for row in rows:
            ws.cell(2, col_idx, f"{row.get('MaxFPS')}FPS")
            ws.cell(3, col_idx, USAGE.get(row.get("Usage"), f"Usage{row.get('Usage')}"))
            fps_groups.setdefault(row.get("MaxFPS"), []).append(col_idx)
            columns.append((col_idx, row))
            col_idx += 1
        end = col_idx - 1
        ws.cell(1, start, PLATFORM.get(platform, f"Platform{platform}"))
        if start < end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        for cols in fps_groups.values():
            if len(cols) > 1:
                ws.merge_cells(start_row=2, start_column=min(cols), end_row=2, end_column=max(cols))
    return columns


def _include(row: dict) -> bool:
    platform = PLATFORM.get(row.get("Platform"), "")
    usage = USAGE.get(row.get("Usage"), str(row.get("Usage")))
    if platform in {"PC", "SteamDeck"}:
        return "PC" in usage
    if platform in {"NSW2", "NSW2_HANDHELD"}:
        return usage in {"FastLoadMode", "DialoguePlaying"} or "Default" in usage
    return "Default" in usage


def _flatten(value: Any, prefix: str = "") -> dict:
    if isinstance(value, dict):
        out = {}
        for key, item in value.items():
            name = f"{prefix}_{key}" if prefix else str(key)
            out.update(_flatten(item, name))
        return out
    if isinstance(value, list):
        return {prefix: str(value)}
    return {prefix: value}


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return str(value)
    return value


def _color_row(ws, row_idx: int, values: list) -> None:
    normalized = [str(v) for v in values if v is not None]
    unique = sorted(set(normalized))
    if len(unique) <= 1:
        return
    for col_idx, value in enumerate(values, start=3):
        if value is None:
            continue
        ratio = unique.index(str(value)) / (len(unique) - 1)
        red = int(255 - 72 * ratio)
        green = int(220 + 28 * ratio)
        blue = int(220 - 80 * ratio)
        ws.cell(row_idx, col_idx).fill = PatternFill(
            start_color=f"{red:02X}{green:02X}{blue:02X}",
            end_color=f"{red:02X}{green:02X}{blue:02X}",
            fill_type="solid",
        )


def _style(ws) -> None:
    ws.freeze_panes = "C4"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row <= 3:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    for row in ws.iter_rows(min_row=4, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
