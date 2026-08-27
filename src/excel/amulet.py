from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.excel.style import apply_rare_style

HEADER_FILL = PatternFill("solid", fgColor="C65911")
ALTERNATE_FILL = PatternFill("solid", fgColor="FFF4EC")
HEADER_BORDER = Border(bottom=Side(style="medium", color="8C4A1F"))
ROW_BORDER = Border(bottom=Side(style="thin", color="E7D8CF"))
RARITY_NUMBER_FORMAT = '"Rare."0'

CENTER_COLUMNS = {
    "Index",
    "AmuletType",
    "Rarity",
    "SkillPt1",
    "SkillPt2",
    "SkillPt3",
    "SlotPt",
    "WeaponSlots",
    "ArmorSlots",
    "SkillPt",
    "SkillId",
    "SkillLevel",
}

MIN_COLUMN_WIDTHS = {
    "AmuletType": 14.0,
    "AmuletName": 20.0,
    "Rarity": 9.0,
    "WeaponSlots": 18.0,
    "ArmorSlots": 18.0,
    "SkillId": 20.0,
    "SkillName": 28.0,
}


def style_amulet_workbook(workbook) -> None:
    for sheet in workbook.worksheets:
        _style_sheet(sheet)


def _style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    sheet.row_dimensions[1].height = 24

    headers = {
        cell.value: cell.column
        for cell in sheet[1]
        if isinstance(cell.value, str)
    }
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = HEADER_BORDER

    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 20
        for cell in sheet[row_index]:
            if row_index % 2 == 0:
                cell.fill = ALTERNATE_FILL
            cell.border = ROW_BORDER
            header = sheet.cell(1, cell.column).value
            horizontal = "center" if header in CENTER_COLUMNS else "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")

    rarity_column = headers.get("Rarity")
    if rarity_column:
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row_index, rarity_column)
            if not isinstance(cell.value, int):
                continue
            cell.number_format = RARITY_NUMBER_FORMAT
            apply_rare_style(cell, cell.value - 1, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ROW_BORDER

    for header, width in MIN_COLUMN_WIDTHS.items():
        column = headers.get(header)
        if not column:
            continue
        letter = get_column_letter(column)
        current = sheet.column_dimensions[letter].width or 0
        sheet.column_dimensions[letter].width = max(current, width)

    if sheet.title == "SkillPool":
        _style_skill_pool(sheet)


def _style_skill_pool(sheet) -> None:
    sheet.auto_filter.ref = None
    sheet.column_dimensions["A"].width = max(
        sheet.column_dimensions["A"].width or 0,
        16.0,
    )
    for column in range(2, sheet.max_column + 1):
        letter = get_column_letter(column)
        current = sheet.column_dimensions[letter].width or 0
        sheet.column_dimensions[letter].width = max(current, 20.0)

    if sheet.max_row < 2:
        return
    sheet.merge_cells(start_row=2, start_column=1, end_row=sheet.max_row, end_column=1)
    label = sheet["A2"]
    label.font = Font(bold=True)
    label.alignment = Alignment(horizontal="center", vertical="center")
    label.fill = ALTERNATE_FILL
    label.border = ROW_BORDER
