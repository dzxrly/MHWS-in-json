from __future__ import annotations

from pathlib import Path
import math
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.converters.action_values import (
    ActionValueWorkbookData, LEADING_COLUMNS, RowGroup,
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
MAPPED_FILLS = (
    PatternFill("solid", fgColor="EAF2F8"),
    PatternFill("solid", fgColor="EDF7ED"),
)
MAPPED_ACTION_FILLS = (
    PatternFill("solid", fgColor="D6E4F0"),
    PatternFill("solid", fgColor="DDEEDC"),
)
UNMAPPED_FILL = PatternFill("solid", fgColor="FCE4D6")
UNMAPPED_ACTION_FILL = PatternFill("solid", fgColor="F4B183")

HEADER_BORDER = Border(bottom=Side(style="medium", color="17365D"))
VERTICAL_SIDE = Side(style="thin", color="D9E2F3")
GROUP_SIDE = Side(style="medium", color="7F8C8D")

CENTER_COLUMNS = {
    "MappingKind",
    "ResourceRole",
    "MappingConfidence",
    "sourceRequestSetOrdinal",
    "requestSetID",
    "groupIndex",
    "status",
    "requestSetIndex",
    "keyHash",
    "KeyNameMMHash",
    "userDataType",
}

FIXED_WIDTHS = {
    "MappingName": 34.0,
    "MappingKind": 12.0,
    "MappingIdentity": 42.0,
    "MappingInternalName": 28.0,
    "MappingNameSource": 24.0,
    "ResourceRole": 16.0,
    "MappingConfidence": 16.0,
    "MappingCondition": 38.0,
    "MappingSource": 36.0,
    "sourceRequestSetOrdinal": 14.0,
    "requestSetID": 14.0,
    "groupIndex": 12.0,
    "status": 10.0,
    "requestSetIndex": 14.0,
    "keyHash": 14.0,
    "KeyNameMMHash": 16.0,
    "name": 14.0,
    "keyName": 18.0,
    "userDataType": 22.0,
}

HAN_RE = re.compile(r"[\u4e00-\u9fff]")


def write_action_value_workbook(
    path: Path,
    data: ActionValueWorkbookData,
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    for sheet_index, (sheet_name, rows) in enumerate(data.sheets.items()):
        sheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
        sheet.title = sheet_name
        columns = data.columns[sheet_name]
        sheet.append([_source_line(data.sources[sheet_name])])
        sheet.append(list(columns))
        for row in rows:
            sheet.append([_safe_cell(row.get(column)) for column in columns])

    style_action_value_workbook(workbook, data)
    workbook.save(path)
    return path


def style_action_value_workbook(
    workbook: Workbook,
    data: ActionValueWorkbookData,
) -> None:
    for sheet in workbook.worksheets:
        _style_sheet(
            sheet,
            data.columns[sheet.title],
            data.groups[sheet.title],
        )


def _style_sheet(
    sheet,
    columns: tuple[str, ...],
    groups: tuple[RowGroup, ...],
) -> None:
    sheet.freeze_panes = f"{get_column_letter(len(LEADING_COLUMNS) + 1)}3"
    name_column = columns.index("MappingName") + 1
    metadata_start = columns.index("MappingKind") + 1
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85
    sheet.sheet_properties.tabColor = "ED7D31" if sheet.title == "Ammo" else "5B9BD5"
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 48

    if sheet.max_column > 1:
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=sheet.max_column,
        )
    source_cell = sheet.cell(1, 1)
    source_cell.fill = PatternFill("solid", fgColor="D9EAF7")
    source_cell.font = Font(italic=True, color="1F4E78")
    source_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    source_cell.border = Border(bottom=Side(style="thin", color="9EADBA"))

    for cell in sheet[2]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = HEADER_BORDER

    wrapped_columns = [
        (columns.index(name) + 1, FIXED_WIDTHS[name] - 2.0)
        for name in ("MappingInternalName", "MappingCondition")
    ]
    for row_index in range(3, sheet.max_row + 1):
        wrapped_lines = max(
            sum(
                max(1, math.ceil(_text_width(line) / width))
                for line in str(sheet.cell(row_index, column).value or "").splitlines()
            )
            for column, width in wrapped_columns
        )
        sheet.row_dimensions[row_index].height = min(
            409.0, max(21.0, 15.0 * wrapped_lines + 6.0)
        )
        for column_index, header in enumerate(columns, start=1):
            cell = sheet.cell(row_index, column_index)
            cell.alignment = Alignment(
                horizontal="center" if header in CENTER_COLUMNS else "left",
                vertical="center",
                wrap_text=header
                in {
                    "MappingName",
                    "MappingIdentity",
                    "MappingInternalName",
                    "MappingCondition",
                    "MappingSource",
                    "name",
                    "keyName",
                }
                # Keep raw text inside its cell before the trailing metadata.
                or column_index == metadata_start - 1,
            )

    mapped_index = 0
    for group in groups:
        if group.unmapped:
            row_fill = UNMAPPED_FILL
            action_fill = UNMAPPED_ACTION_FILL
        else:
            color_index = mapped_index % len(MAPPED_FILLS)
            row_fill = MAPPED_FILLS[color_index]
            action_fill = MAPPED_ACTION_FILLS[color_index]
            mapped_index += 1
        _style_group(sheet, group, row_fill, action_fill, name_column)

        if not group.unmapped and group.end_row > group.start_row:
            sheet.merge_cells(
                start_row=group.start_row,
                start_column=name_column,
                end_row=group.end_row,
                end_column=name_column,
            )
        action_cell = sheet.cell(group.start_row, name_column)
        action_cell.font = Font(bold=True, color="9C5700" if group.unmapped else "1F1F1F")
        action_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    _set_column_widths(sheet, columns)


def _style_group(
    sheet, group: RowGroup, row_fill: PatternFill, action_fill: PatternFill, name_column: int,
) -> None:
    for row_index in range(group.start_row, group.end_row + 1):
        top = GROUP_SIDE if row_index == group.start_row else Side(style=None)
        bottom = GROUP_SIDE if row_index == group.end_row else Side(style=None)
        for column_index in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_index, column_index)
            cell.fill = action_fill if column_index == name_column else row_fill
            cell.border = Border(
                left=VERTICAL_SIDE,
                right=VERTICAL_SIDE,
                top=top,
                bottom=bottom,
            )


def _set_column_widths(sheet, columns: tuple[str, ...]) -> None:
    for column_index, header in enumerate(columns, start=1):
        letter = get_column_letter(column_index)
        if header in FIXED_WIDTHS:
            sheet.column_dimensions[letter].width = FIXED_WIDTHS[header]
            continue

        width = max(10.0, _text_width(header) + 2.0)
        for row_index in range(3, min(sheet.max_row, 251) + 1):
            width = max(width, _text_width(sheet.cell(row_index, column_index).value) + 2.0)
        sheet.column_dimensions[letter].width = min(width, 24.0)


def _safe_cell(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        value = str(value)
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def _source_line(sources: tuple[str, ...]) -> str:
    if not sources:
        return "RCOL sources: [none]"
    return f"RCOL sources ({len(sources)}): " + " | ".join(sources)


def _text_width(value: Any) -> float:
    if value is None:
        return 0.0
    total = 0.0
    for char in str(value):
        if char.isascii() and char.isalnum():
            total += 1.2
        elif HAN_RE.search(char):
            total += 2.0
        else:
            total += 1.0
    return total
