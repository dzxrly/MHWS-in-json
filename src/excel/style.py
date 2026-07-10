import re
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HAN_RE = re.compile(r"[\u4e00-\u9fff]")
RARE_RE = re.compile(r"^RARE(\d+)$")
RARE_COLORS = {
    0: "969696",
    1: "DEDEDE",
    2: "A4C43B",
    3: "47A33F",
    4: "5CAEBB",
    5: "575FD9",
    6: "9272E3",
    7: "C76D46",
    8: "B3436A",
    9: "2EC9E6",
    10: "F2C21D",
    11: "B4F5FF",
}


def style_workbook(workbook, max_width: float = 80.0) -> None:
    for sheet in workbook.worksheets:
        _style_header(sheet)
        _align_cells(sheet)
        _fit_columns(sheet, max_width)
        _wrap_text_columns(sheet, {"Explain", "RawExplain"})
        _apply_rare_colors(sheet)


def _style_header(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _align_cells(sheet) -> None:
    alignment = Alignment(horizontal="left", vertical="center")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment


def _fit_columns(sheet, max_width: float) -> None:
    for col in range(1, sheet.max_column + 1):
        width = 6.0
        for row in range(1, sheet.max_row + 1):
            width = max(width, _text_width(sheet.cell(row, col).value))
        sheet.column_dimensions[get_column_letter(col)].width = min(width, max_width)


def _wrap_text_columns(sheet, names: set[str]) -> None:
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(1, col).value in names:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, col).alignment = wrap


def _apply_rare_colors(sheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            match = RARE_RE.match(cell.value)
            if not match:
                continue
            rare = int(match.group(1))
            color = _blend_with_white(RARE_COLORS.get(rare, "FFFFFF"), 0.5)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(color="FFFFFF" if _brightness(color) < 64 else "000000")
            cell.value = rare + 1


def _text_width(value: Any) -> float:
    if value is None:
        return 0.0
    total = 0.0
    for char in str(value):
        if char.isascii() and char.isalnum():
            total += 1.2
        elif HAN_RE.search(char):
            total += 2.0
        else:
            total += 1.0
    return total


def _brightness(color: str) -> float:
    r, g, b = int(color[:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    return (r * 299 + g * 587 + b * 114) / 1000


def _blend_with_white(color: str, opacity: float) -> str:
    rgb = [int(color[i : i + 2], 16) for i in (0, 2, 4)]
    return "".join(f"{int(c * opacity + 255 * (1 - opacity)):02X}" for c in rgb)
