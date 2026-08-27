import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from config import LANGUAGES
from src.converters.amulet import (
    PATHS,
    build_amulet_pools,
    build_amulet_workbook_sheets,
    export_amulet_pools,
    load_amulet_catalog,
)
from src.excel.amulet import RARITY_NUMBER_FORMAT, style_amulet_workbook
from src.excel.writer import write_workbook


class AmuletTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tables = _tables()
        self.catalog = load_amulet_catalog(
            lambda path: self.tables[_key_for_path(path)]
        )
        names = {
            "amulet-guid": "Localized Amulet",
            "skill-guid-1": "Attack",
            "skill-guid-2": "Defense",
        }
        self.sheets = build_amulet_workbook_sheets(
            self.catalog,
            lambda guid: names.get(guid or "", ""),
        )

    def test_workbook_rows_split_slots_and_display_rarity(self) -> None:
        rows = self.sheets["AmuletPool"]

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["AmuletName"], "Localized Amulet")
        self.assertEqual(rows[0]["Rarity"], 8)
        self.assertEqual(rows[0]["WeaponSlots"], "Lv.1")
        self.assertEqual(rows[0]["ArmorSlots"], "Lv.2, Lv.1")
        self.assertEqual(rows[1]["WeaponSlots"], "")
        self.assertEqual(rows[1]["ArmorSlots"], "Lv.1")
        self.assertNotIn("Lv.0", str(rows))

    def test_message_language_codes_match_game_enum_indices(self) -> None:
        self.assertEqual(LANGUAGES[26].code, "th-TH")
        self.assertEqual(LANGUAGES[32].code, "es-419")

    def test_workbook_uses_normalized_skill_and_slot_tables(self) -> None:
        self.assertEqual(
            self.sheets["SkillPool"],
            [
                {
                    "SkillPt": "Skill / Level",
                    4: "Attack Lv.2",
                    6: "Defense Lv.1",
                },
                {
                    "SkillPt": None,
                    4: "Defense Lv.3",
                    6: None,
                },
            ],
        )
        self.assertEqual(
            self.sheets["SlotPool"],
            [
                {"SlotPt": 1, "WeaponSlots": "", "ArmorSlots": "Lv.1"},
                {
                    "SlotPt": 3,
                    "WeaponSlots": "",
                    "ArmorSlots": "Lv.1, Lv.1, Lv.1",
                },
                {
                    "SlotPt": 12,
                    "WeaponSlots": "Lv.1",
                    "ArmorSlots": "Lv.2, Lv.1",
                },
            ],
        )

    def test_unknown_amulet_type_is_skipped_consistently(self) -> None:
        self.assertEqual(self.catalog.missing_amulet_types, ("AT_MISSING",))
        self.assertEqual(self.catalog.missing_amulet_rows, 1)
        self.assertEqual(
            [row["Index"] for row in self.sheets["AmuletPool"]],
            [7, 8],
        )

        _, amulet_pool = build_amulet_pools(self.catalog, lambda guid: guid or "")
        self.assertEqual([row["id"] for row in amulet_pool], ["7", "8"])

    def test_json_export_keeps_existing_shape_and_rare_literal(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_dir = Path(temp_dir)
            export_amulet_pools(
                output_dir,
                self.catalog,
                lambda guid: f"name:{guid}" if guid else "",
            )
            with (output_dir / "amulet_pool.json").open(encoding="utf-8") as file:
                amulet_pool = json.load(file)

        self.assertEqual(amulet_pool[0]["rare"]["id"], "AT_1")
        self.assertEqual(amulet_pool[0]["rare"]["name"], "name:amulet-guid")
        self.assertEqual(amulet_pool[0]["rare"]["rare"], "Rare 7")
        self.assertEqual(amulet_pool[0]["slot"]["weaponSlot"], [1, 0, 0])
        self.assertEqual(amulet_pool[0]["slot"]["equipmentSlot"], [2, 1, 0])

    def test_workbook_style_formats_rarity_and_navigation(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "AmuletCollection.xlsx"
            write_workbook(path, self.sheets, 80.0, style_amulet_workbook)
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                sheet = workbook["AmuletPool"]
                self.assertEqual(sheet["D2"].value, 8)
                self.assertEqual(sheet["D2"].number_format, RARITY_NUMBER_FORMAT)
                self.assertEqual(sheet.freeze_panes, "A2")
                self.assertEqual(sheet.auto_filter.ref, "A1:J3")
                self.assertFalse(sheet.sheet_view.showGridLines)
                self.assertEqual(sheet["I2"].value, "Lv.1")
                self.assertEqual(sheet["J2"].value, "Lv.2, Lv.1")

                skill_sheet = workbook["SkillPool"]
                self.assertEqual(skill_sheet.max_row, 3)
                self.assertEqual(skill_sheet.max_column, 3)
                self.assertEqual(skill_sheet["A1"].value, "SkillPt")
                self.assertEqual(skill_sheet["B1"].value, 4)
                self.assertEqual(skill_sheet["C1"].value, 6)
                self.assertEqual(skill_sheet["A2"].value, "Skill / Level")
                self.assertEqual(skill_sheet["B2"].value, "Attack Lv.2")
                self.assertEqual(skill_sheet["B3"].value, "Defense Lv.3")
                self.assertEqual(skill_sheet["C3"].value, None)
                self.assertIn("A2:A3", skill_sheet.merged_cells)
                self.assertIsNone(skill_sheet.auto_filter.ref)
            finally:
                workbook.close()


def _key_for_path(path: str) -> str:
    return next(key for key, configured_path in PATHS.items() if configured_path == path)


def _tables() -> dict[str, list[dict]]:
    return {
        "skill_lot": [
            {
                "SkillType": "SKILL_1",
                "SkillLv": 2,
                "SkillPt": 4,
            },
            {
                "SkillType": "SKILL_2",
                "SkillLv": 3,
                "SkillPt": 4,
            },
            {
                "SkillType": "SKILL_2",
                "SkillLv": 1,
                "SkillPt": 6,
            },
        ],
        "pt": [
            {
                "Index": 7,
                "AmuletType": "AT_1",
                "SkillPt_01": 4,
                "SkillPt_02": 6,
                "SkillPt_03": 0,
                "SlotPt": 12,
            },
            {
                "Index": 8,
                "AmuletType": "AT_1",
                "SkillPt_01": 4,
                "SkillPt_02": 0,
                "SkillPt_03": 0,
                "SlotPt": 1,
            },
            {
                "Index": 9,
                "AmuletType": "AT_MISSING",
                "SkillPt_01": 4,
                "SkillPt_02": 0,
                "SkillPt_03": 0,
                "SlotPt": 1,
            },
        ],
        "slot": [
            {
                "SlotPt": 12,
                "SlotType01": "ACC_TYPE_00",
                "SlotLevel01": "Lv1",
                "SlotType02": "ACC_TYPE_01",
                "SlotLevel02": "Lv2",
                "SlotType03": "ACC_TYPE_01",
                "SlotLevel03": "Lv1",
            },
            {
                "SlotPt": 1,
                "SlotType01": "ACC_TYPE_01",
                "SlotLevel01": "Lv1",
                "SlotType02": "ACC_TYPE_01",
                "SlotLevel02": "NONE",
                "SlotType03": "ACC_TYPE_01",
                "SlotLevel03": "NONE",
            },
            {
                "SlotPt": 3,
                "SlotType01": "ACC_TYPE_01",
                "SlotLevel01": "Lv1",
                "SlotType02": "ACC_TYPE_01",
                "SlotLevel02": "Lv1",
                "SlotType03": "ACC_TYPE_01",
                "SlotLevel03": "Lv1",
            },
        ],
        "skill": [
            {"skillId": "SKILL_1", "skillName": "skill-guid-1"},
            {"skillId": "SKILL_2", "skillName": "skill-guid-2"},
        ],
        "amulet": [
            {
                "AmuletType": "AT_1",
                "Name": "amulet-guid",
                "Rare": "RARE7",
            }
        ],
    }


if __name__ == "__main__":
    unittest.main()
