import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from config import ACTION_MAP_PATH, NATIVES_DIR
from src.converters.action_values import (
    ActionValueCatalog,
    MappingBinding,
    SHEET_NAMES,
    build_action_value_workbook,
    load_action_value_catalog,
)
from src.data.action_map import ACTION_MAP_FORMAT, load_action_map
from src.data.rcol import (
    RequestSetKey,
    RequestSetRecord,
    flatten_request_set,
)
from src.excel.action_values import write_action_value_workbook


class ActionValueTests(unittest.TestCase):
    def test_flatten_request_set_keeps_leaf_paths_and_excludes_collider_payload(self) -> None:
        request_set = {
            "requestSetID": 7,
            "keyHash": 99,
            "nativeShapeColliders": [
                {
                    "app.col_user_data.AttackParamPl": {
                        "_Attack": 999.0,
                    }
                }
            ],
            "userData": {
                "app.col_user_data.AttackParamPl": {
                    "_Attack": 81.0,
                    "_ActionTypeFixed": {
                        "app.HitDef.ACTION_TYPE_Serializable": {
                            "_Value": "[1]SLASH",
                        }
                    },
                    "_DisableHitEffect": {
                        "ace.Bitset`1<app.HitDef.HIT_EFFECT_DISABLE_TYPE>": {
                            "_Value": [],
                            "_MaxElement": 3,
                        }
                    },
                    "_EmptyCurve": {"": {}},
                }
            },
        }

        user_data_type, flattened = flatten_request_set(request_set) or ("", {})

        self.assertEqual(user_data_type, "app.col_user_data.AttackParamPl")
        self.assertEqual(flattened["_Attack"], 81.0)
        self.assertEqual(flattened["_ActionTypeFixed._Value"], "[1]SLASH")
        self.assertEqual(flattened["_DisableHitEffect._Value"], "[]")
        self.assertEqual(flattened["_DisableHitEffect._MaxElement"], 3)
        self.assertEqual(flattened["_EmptyCurve"], "{}")
        self.assertFalse(
            any(key.startswith("nativeShapeColliders") for key in flattened)
        )

    def test_flatten_request_set_rejects_technical_reference(self) -> None:
        self.assertIsNone(
            flatten_request_set(
                {
                    "requestSetID": 1,
                    "userData": {"Ref": {"ref_instance_id": None}},
                }
            )
        )

    def test_many_to_many_rows_are_grouped_and_unmapped_rows_are_last(self) -> None:
        record_a = _record(0, 100, 0, "=guard")
        record_b = _record(1, 101, 1, "safe")
        record_unmapped = _record(2, 102, 2, "unknown")
        action_a = MappingBinding(
            identity="action:a",
            scope="Wp00",
            order=(0, 1, 1, "action:a"),
            kind="Action",
            fallback_name="Action A",
            internal_name="ACTION_A",
            source="test",
        )
        action_b = MappingBinding(
            identity="action:b",
            scope="Wp00",
            order=(0, 2, 2, "action:b"),
            kind="Action",
            fallback_name="Action B",
            internal_name="ACTION_B",
            source="test",
        )
        catalog = ActionValueCatalog(
            records={"Wp00": (record_a, record_b, record_unmapped)},
            actions={},
            bindings={
                record_a.key: (action_a, action_b),
                record_b.key: (action_a,),
            },
            mapping_source_counts={"test": 3},
        )

        data = build_action_value_workbook(catalog, lambda _guid: "")
        rows = data.sheets["Wp00_LongSword"]

        self.assertEqual(
            [(row["MappingName"], row["requestSetID"]) for row in rows],
            [
                ("Action A", 0),
                ("Action A", 1),
                ("Action B", 0),
                (None, 2),
            ],
        )
        self.assertEqual(
            [
                (group.start_row, group.end_row, group.unmapped)
                for group in data.groups["Wp00_LongSword"]
            ],
            [(3, 4, False), (5, 5, False), (6, 6, True)],
        )
        self.assertEqual(
            data.sources["Wp00_LongSword"],
            ("Wp00/Collision/Collider/Wp00_Attack.rcol.38.json",),
        )
        self.assertNotIn("rcol", data.columns["Wp00_LongSword"])
        audit = data.audits["Wp00_LongSword"]
        self.assertEqual(audit.eligible_request_sets, 3)
        self.assertEqual(audit.mapped_request_sets, 2)
        self.assertEqual(audit.unmapped_request_sets, 1)
        self.assertEqual(audit.displayed_rows, 4)

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "WeaponActionValues.xlsx"
            write_action_value_workbook(path, data)
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                self.assertEqual(workbook.sheetnames, list(SHEET_NAMES.values()))
                sheet = workbook["Wp00_LongSword"]
                self.assertEqual(sheet.freeze_panes, "B3")
                self.assertFalse(sheet.sheet_view.showGridLines)
                self.assertTrue(
                    any(str(cell_range).startswith("A1:") for cell_range in sheet.merged_cells)
                )
                self.assertIn("A3:A4", sheet.merged_cells)
                self.assertNotIn("A5:A6", sheet.merged_cells)
                self.assertTrue(sheet["A1"].value.startswith("RCOL sources (1): "))
                self.assertEqual(sheet["A2"].value, "MappingName")
                self.assertEqual(sheet["B3"].value, "Action")
                self.assertIsNone(sheet["A6"].value)
                self.assertEqual(sheet["B3"].border.bottom.style, None)
                self.assertEqual(sheet["B4"].border.bottom.style, "medium")
                self.assertEqual(sheet["A6"].fill.fgColor.rgb, "00F4B183")

                note_column = data.columns["Wp00_LongSword"].index("note") + 1
                self.assertEqual(sheet.cell(3, note_column).value, "'=guard")
            finally:
                workbook.close()

    def test_action_map_is_the_only_binding_authority_and_keeps_many_to_many(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            natives_dir = _synthetic_natives(root, [(0, 100), (1, 101), (2, 102)])
            rcol = "Wp00/Collision/Wp00_Test.rcol.38.json"
            action_map = root / "ActionMap.json"
            _write_action_map(
                action_map,
                [
                    _relation("guide:Wp00:10", "guid-a", 10, rcol, 0, 100, 0),
                    _relation("guide:Wp00:20", "guid-b", 20, rcol, 0, 100, 0),
                    _relation("guide:Wp00:10", "guid-a", 10, rcol, 1, 101, 1),
                    {
                        **_relation("unnamed", "", 30, rcol, 2, 102, 2),
                        "actionJapaneseName": "",
                    },
                ],
            )

            catalog = load_action_value_catalog(natives_dir, action_map)
            data = build_action_value_workbook(
                catalog,
                lambda guid: {"guid-a": " 同名动作\t", "guid-b": "同名动作"}.get(
                    guid,
                    "",
                ),
            )

        records = catalog.records["Wp00"]
        self.assertEqual(len(catalog.bindings[records[0].key]), 2)
        self.assertEqual(len(catalog.bindings[records[1].key]), 1)
        self.assertEqual(len(catalog.bindings[records[2].key]), 1)
        self.assertTrue(
            all(
                source.startswith("action_map:")
                for source in catalog.mapping_source_counts
            )
        )
        rows = data.sheets["Wp00_LongSword"]
        self.assertEqual(
            [(row["MappingName"], row["requestSetID"]) for row in rows],
            [
                ("同名动作", 0),
                ("同名动作", 1),
                ("同名动作", 0),
                ("Action 30", 2),
            ],
        )
        groups = data.groups["Wp00_LongSword"]
        self.assertEqual(
            [(group.start_row, group.end_row, group.unmapped) for group in groups],
            [(3, 4, False), (5, 5, False), (6, 6, False)],
        )
        assert catalog.action_map_audit is not None
        self.assertEqual(catalog.action_map_audit.relations, 4)
        self.assertEqual(catalog.action_map_audit.unnamed_relations, 1)
        self.assertEqual(catalog.action_map_audit.bound_request_sets, 3)

    def test_resource_relation_uses_fallback_name_and_keeps_condition(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            natives_dir = _synthetic_natives(root, [(7, 107)])
            rcol = "Wp00/Collision/Wp00_Test.rcol.38.json"
            action_map = root / "ActionMap.json"
            _write_action_map(
                action_map,
                [],
                [
                    _resource_relation(
                        "shell:Wp00:7:variant:test",
                        rcol,
                        7,
                        107,
                        0,
                    )
                ],
            )

            catalog = load_action_value_catalog(natives_dir, action_map)
            data = build_action_value_workbook(catalog, lambda _guid: "")

        row = data.sheets["Wp00_LongSword"][0]
        self.assertEqual(row["MappingName"], "SPECIAL_SHELL [TEST_MODE]")
        self.assertEqual(row["MappingKind"], "Resource")
        self.assertEqual(row["MappingIdentity"], "shell:Wp00:7:variant:test")
        self.assertEqual(row["MappingInternalName"], "SPECIAL_SHELL_TEST")
        self.assertEqual(row["ResourceRole"], "shell")
        self.assertEqual(row["MappingConfidence"], "derived")
        self.assertIn('"value":"TEST_MODE"', row["MappingCondition"])
        assert catalog.action_map_audit is not None
        self.assertEqual(catalog.action_map_audit.action_relations, 0)
        self.assertEqual(catalog.action_map_audit.resource_relations, 1)
        self.assertEqual(catalog.action_map_audit.bound_request_sets, 1)

    def test_action_map_rejects_stale_or_ambiguous_exact_identity(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            natives_dir = _synthetic_natives(root, [(0, 100)])
            action_map = root / "ActionMap.json"
            _write_action_map(
                action_map,
                [
                    _relation(
                        "guide:Wp00:10",
                        "guid-a",
                        10,
                        "Wp00/Collision/Wp00_Test.rcol.38.json",
                        0,
                        999,
                        0,
                    )
                ],
            )
            with self.assertRaisesRegex(ValueError, "exact current requestSet"):
                load_action_value_catalog(natives_dir, action_map)

    def test_action_map_rejects_v1_format(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "ActionMap.json"
            path.write_text(
                json.dumps(
                    {
                        "_format": "mhws_static_action_request_set_map_v1",
                        "relations": [],
                    }
                ),
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "Unsupported ActionMap"):
                load_action_map(path)

    def test_real_source_contract_has_all_scopes_and_unmapped_tail_blocks(self) -> None:
        catalog = load_action_value_catalog(NATIVES_DIR, ACTION_MAP_PATH)
        data = build_action_value_workbook(catalog, lambda _guid: "")

        self.assertEqual(set(catalog.records), set(SHEET_NAMES))
        self.assertGreater(sum(len(rows) for rows in catalog.records.values()), 0)
        self.assertGreater(sum(catalog.mapping_source_counts.values()), 0)
        self.assertTrue(catalog.records["Ammo"])
        assert catalog.action_map_audit is not None
        self.assertGreater(catalog.action_map_audit.named_relations, 0)

        for sheet_name, rows in data.sheets.items():
            groups = data.groups[sheet_name]
            unmapped_groups = [group for group in groups if group.unmapped]
            self.assertLessEqual(len(unmapped_groups), 1)
            if unmapped_groups:
                unmapped_group = unmapped_groups[0]
                self.assertIs(unmapped_group, groups[-1])
                unmapped_start = unmapped_group.start_row - 3
                self.assertTrue(
                    all(row["MappingName"] for row in rows[:unmapped_start])
                )
                self.assertTrue(
                    all(
                        row["MappingName"] is None
                        for row in rows[unmapped_start:]
                    )
                )

            expected_columns = data.columns[sheet_name]
            self.assertEqual(expected_columns[0], "MappingName")
            self.assertEqual(expected_columns[1], "MappingKind")
            self.assertIn("_ActionTypeFixed._Value", expected_columns)


def _record(
    request_set_id: int,
    key_hash: int,
    source_ordinal: int,
    note: str,
) -> RequestSetRecord:
    key = RequestSetKey(
        scope="Wp00",
        rcol="Wp00/Collision/Collider/Wp00_Attack.rcol.38.json",
        request_set_id=request_set_id,
        key_hash=key_hash,
        source_ordinal=source_ordinal,
    )
    return RequestSetRecord(
        key=key,
        user_data_type="app.col_user_data.AttackParamPl",
        properties={
            "sourceRequestSetOrdinal": source_ordinal,
            "requestSetID": request_set_id,
            "keyHash": key_hash,
            "userDataType": "AttackParamPl",
            "_Attack": 10.0 + request_set_id,
            "_ActionTypeFixed._Value": "[1]SLASH",
            "note": note,
        },
    )


def _synthetic_natives(
    root: Path,
    rows: list[tuple[int, int]],
) -> Path:
    natives_dir = root / "natives"
    path = (
        natives_dir
        / "STM"
        / "GameDesign"
        / "Player"
        / "ActionData"
        / "Wp00"
        / "Collision"
        / "Wp00_Test.rcol.38.json"
    )
    path.parent.mkdir(parents=True)
    path.write_text(
        json.dumps(
            {
                "requestSets": [
                    {
                        "requestSetID": request_set_id,
                        "keyHash": key_hash,
                        "userData": {
                            "app.col_user_data.AttackParamPl": {
                                "_Attack": float(request_set_id),
                                "_ActionTypeFixed": {
                                    "app.HitDef.ACTION_TYPE_Serializable": {
                                        "_Value": "[1]SLASH",
                                    }
                                },
                            }
                        },
                    }
                    for request_set_id, key_hash in rows
                ]
            }
        ),
        encoding="utf-8",
    )
    return natives_dir


def _relation(
    identity: str,
    guid: str,
    guide_id: int,
    rcol: str,
    request_set_id: int,
    key_hash: int,
    source_ordinal: int,
) -> dict[str, object]:
    return {
        "scope": "Wp00",
        "actionIdentity": identity,
        "actionGuideId": guide_id,
        "actionOrder": guide_id,
        "actionInternalName": identity,
        "actionNameGuid": guid,
        "actionJapaneseName": f"动作 {guide_id}" if guid else "",
        "fallbackName": f"Action {guide_id}",
        "source": "test_static",
        "resolutionMethods": ["test_exact"],
        "confidence": "proven",
        "conditions": [],
        "rcol": rcol,
        "requestSetId": request_set_id,
        "keyHash": key_hash,
        "sourceRequestSetOrdinal": source_ordinal,
    }


def _resource_relation(
    identity: str,
    rcol: str,
    request_set_id: int,
    key_hash: int,
    source_ordinal: int,
) -> dict[str, object]:
    return {
        "scope": "Wp00",
        "resourceIdentity": identity,
        "resourceOrder": 10,
        "resourceInternalName": "SPECIAL_SHELL_TEST",
        "resourceNameGuid": "",
        "resourceJapaneseName": "",
        "resourceNameSource": "shell_fixed_enum",
        "resourceNameSuffix": " [TEST_MODE]",
        "fallbackName": "SPECIAL_SHELL",
        "resourceRole": "shell",
        "source": "test_resource",
        "resolutionMethods": ["test_exact"],
        "confidence": "derived",
        "conditions": [
            {"kind": "request_set_resource_marker", "value": "TEST_MODE"}
        ],
        "rcol": rcol,
        "requestSetId": request_set_id,
        "keyHash": key_hash,
        "sourceRequestSetOrdinal": source_ordinal,
    }


def _write_action_map(
    path: Path,
    action_relations: list[dict[str, object]],
    resource_relations: list[dict[str, object]] | None = None,
) -> None:
    path.write_text(
        json.dumps(
            {
                "_format": ACTION_MAP_FORMAT,
                "actionRelations": action_relations,
                "resourceRelations": resource_relations or [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


if __name__ == "__main__":
    unittest.main()
